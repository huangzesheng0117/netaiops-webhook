import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from urllib.parse import unquote, urlparse, parse_qs

import yaml
from openpyxl import load_workbook

BASE_DIR = Path("/opt/netaiops-webhook")
INPUT_DIR = BASE_DIR / "input" / "monitoring"
OUT_DIR = BASE_DIR / "catalogs" / "three_layer"

def read_json(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_lower(v):
    if v is None:
        return ""
    return str(v).strip().lower()

def split_multi(v: str):
    if not v:
        return []
    parts = re.split(r"[|,]", v)
    return [p.strip() for p in parts if p.strip()]

def extract_list_from_query(query: str, key: str):
    if not query:
        return []
    # ip=~"a|b" / ip="a"
    out = []
    for m in re.finditer(rf'{re.escape(key)}\s*=~?\s*"([^"]+)"', query):
        out.extend(split_multi(m.group(1)))
    # key='...'
    for m in re.finditer(rf"{re.escape(key)}\s*=~?\s*'([^']+)'", query):
        out.extend(split_multi(m.group(1)))
    # de-dup keep order
    seen = set()
    result = []
    for x in out:
        if x not in seen:
            seen.add(x)
            result.append(x)
    return result

def infer_direction(name: str, desc: str):
    text = f"{name} {desc}"
    if "入向" in text or "inbound" in safe_lower(text):
        return "inbound"
    if "出向" in text or "outbound" in safe_lower(text):
        return "outbound"
    return ""

def infer_carrier(name: str, desc: str):
    text = f"{name} {desc}"
    for carrier in ["电信", "联通", "移动", "鹏博士", "BGP", "互联网"]:
        if carrier in text:
            return carrier
    return ""

def infer_link_name(name: str):
    text = str(name or "")
    for sep in ["利用率", "丢包", "错包", "BGP", "状态", "_100M", "_200M", "-入向", "-出向"]:
        if sep in text:
            text = text.split(sep)[0]
            break
    return text.strip("_- ")

def detect_family(name: str, query: str, desc: str):
    text = f"{name} {query} {desc}".lower()
    if any(k in text for k in ["bgp邻居", "ospf邻居", "bfd邻居"]):
        return "routing_neighbor_down"
    if "光功率" in f"{name}{desc}":
        return "optical_power_abnormal"
    if any(k in f"{name}{desc}" for k in ["丢包", "错包", "discard", "crc", "FC端口错误"]):
        return "interface_packet_loss_or_discards_high"
    if any(k in f"{name}{desc}" for k in ["利用率", "带宽"]):
        return "interface_or_link_utilization_high"
    if any(k in f"{name}{desc}" for k in ["连接数", "会话"]):
        return "connection_or_session_anomaly"
    if any(k in f"{name}{desc}" for k in ["风扇", "电源", "温度", "板卡", "机框", "主板", "处理器", "存储", "Flash"]):
        return "hardware_component_abnormal"
    if any(k in f"{name}{desc}" for k in ["CPU", "cpu"]):
        return "device_cpu_high"
    if any(k in f"{name}{desc}" for k in ["内存", "TMM"]):
        return "device_memory_high"
    if any(k in f"{name}{desc}" for k in ["DNS", "解析率", "请求率"]):
        return "dns_quality_or_traffic_anomaly"
    if any(k in f"{name}{desc}" for k in ["down", "flap", "状态"]) and ("interface" in text or "端口" in f"{name}{desc}" or "OperState" in query):
        return "interface_status_or_flap"
    return "generic_network_readonly"

def infer_vendor_hint(name: str, query: str, file_name: str, sample_jobs):
    text = f"{name} {query} {file_name} {' '.join(sample_jobs)}".lower()
    if "fortigate" in text or "fg" in text:
        return "fortigate"
    if "hillstone" in text:
        return "hillstone"
    if "f5" in text or "ltm" in text or "gtm" in text or "dns-f5" in text:
        return "f5"
    if "huawei" in text:
        return "huawei"
    if "h3c" in text or "comware" in text:
        return "h3c"
    if "cimc" in text:
        return "cimc"
    if any(k in text for k in ["cisco", "nxos", "catalyst", "sw-cisco", "ios-xe"]):
        return "cisco"
    return "generic_network"

def load_inventory(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(len(headers))}
        device_ip = str(item.get("hostname") or "").strip()
        rows.append({
            "site": str(item.get("idc") or "").strip(),
            "hostname": str(item.get("sysname") or "").strip(),
            "device_ip": device_ip,
            "vendor": str(item.get("vendor") or "").strip(),
            "platform": str(item.get("model") or item.get("os_family") or "").strip(),
            "os_family": str(item.get("os_family") or "").strip(),
            "role": str(item.get("role") or "").strip(),
            "ssh_port": str(item.get("ssh_port") or "22").strip(),
        })
    return rows

def build():
    rules_json = read_json(INPUT_DIR / "rules.txt")
    alerts_json = read_json(INPUT_DIR / "alerts.txt")
    targets_json = read_json(INPUT_DIR / "targets.txt")

    inventory_rows = load_inventory(INPUT_DIR / "device_inventory.xlsx")
    inventory_by_ip = {x["device_ip"]: x for x in inventory_rows if x["device_ip"]}

    target_rows = []
    for t in targets_json.get("data", {}).get("activeTargets", []):
        labels = t.get("labels", {}) or {}
        ip = str(labels.get("ip") or "").strip()
        job = str(labels.get("job") or "").strip()
        instance = str(labels.get("instance") or "").strip()
        target_rows.append({
            "ip": ip,
            "job": job,
            "instance": instance,
            "health": t.get("health", ""),
            "sysname": inventory_by_ip.get(ip, {}).get("hostname", ""),
            "vendor": inventory_by_ip.get(ip, {}).get("vendor", ""),
            "platform": inventory_by_ip.get(ip, {}).get("platform", ""),
            "site": inventory_by_ip.get(ip, {}).get("site", ""),
        })

    alerts_by_name = defaultdict(list)
    for a in alerts_json.get("data", {}).get("alerts", []):
        alerts_by_name[str(a.get("name") or "")].append(a)

    layer1 = {
        "layer1_alert_families": {
            "interface_or_link_utilization_high": {
                "layer1_name": "接口/链路利用率过高",
                "playbook_seed": "interface_or_link_utilization_high",
            },
            "interface_packet_loss_or_discards_high": {
                "layer1_name": "接口丢包/错包/丢弃异常",
                "playbook_seed": "interface_packet_loss_or_discards_high",
            },
            "interface_status_or_flap": {
                "layer1_name": "接口状态/抖动异常",
                "playbook_seed": "interface_status_or_flap",
            },
            "routing_neighbor_down": {
                "layer1_name": "路由邻居异常",
                "playbook_seed": "routing_neighbor_down",
            },
            "device_cpu_high": {
                "layer1_name": "CPU 利用率过高",
                "playbook_seed": "device_cpu_high",
            },
            "device_memory_high": {
                "layer1_name": "内存利用率过高",
                "playbook_seed": "device_memory_high",
            },
            "hardware_component_abnormal": {
                "layer1_name": "硬件部件异常",
                "playbook_seed": "hardware_component_abnormal",
            },
            "optical_power_abnormal": {
                "layer1_name": "光功率异常",
                "playbook_seed": "optical_power_abnormal",
            },
            "connection_or_session_anomaly": {
                "layer1_name": "连接数/会话异常",
                "playbook_seed": "connection_or_session_anomaly",
            },
            "dns_quality_or_traffic_anomaly": {
                "layer1_name": "DNS 质量/流量异常",
                "playbook_seed": "dns_quality_or_traffic_anomaly",
            },
            "generic_network_readonly": {
                "layer1_name": "通用只读排障",
                "playbook_seed": "generic_network_readonly",
            },
        }
    }

    layer2 = []
    layer3 = []

    for group in rules_json.get("data", {}).get("groups", []):
        for rule in group.get("rules", []):
            name = str(rule.get("name") or "").strip()
            query = str(rule.get("query") or "").strip()
            file_name = os.path.basename(str(rule.get("file") or ""))
            desc = str((rule.get("annotations") or {}).get("description") or "").strip()
            severity = str((rule.get("labels") or {}).get("severity") or "").strip()

            sample_alerts = alerts_by_name.get(name, [])
            sample_jobs = []
            sample_keys = set()
            sample_label_values = {}
            for a in sample_alerts[:5]:
                labels = a.get("labels", {}) or {}
                sample_jobs.append(str(labels.get("job") or ""))
                for k,v in labels.items():
                    sample_keys.add(k)
                    if k not in sample_label_values and v not in (None, ""):
                        sample_label_values[k] = str(v)

            family = detect_family(name, query, desc)
            vendor_hint = infer_vendor_hint(name, query, file_name, sample_jobs)
            static_extract = {
                "ip_matchers": extract_list_from_query(query, "ip"),
                "ifname_matchers": extract_list_from_query(query, "ifName"),
                "instance_matchers": extract_list_from_query(query, "instance"),
                "job_matchers": extract_list_from_query(query, "job"),
            }
            runtime_templates = [k for k in ["ip","ifName","ifAlias","instance","job","monitor","sysName","hostname","vendor"] if k in sample_keys]
            layer2.append({
                "rule_id": f"{group.get('name','')}::{(rule.get('labels') or {}).get('group','')}::{name}",
                "match": {"alertname": name},
                "source_rule_file": file_name,
                "vendor_hint": vendor_hint,
                "family": family,
                "playbook_type": family,
                "severity": severity,
                "job_patterns": [x for x in (static_extract["job_matchers"] + sample_jobs) if x],
                "notes": desc,
                "runtime_label_templates": runtime_templates,
            })

            layer3.append({
                "rule_id": f"{group.get('name','')}::{(rule.get('labels') or {}).get('group','')}::{name}",
                "alertname": name,
                "source_rule_file": file_name,
                "vendor_hint": vendor_hint,
                "family": family,
                "enrichment_strategy": "hybrid_rule_expr_plus_runtime_labels",
                "static_extract": static_extract,
                "runtime_label_templates": runtime_templates,
                "direction": infer_direction(name, desc),
                "carrier": infer_carrier(name, desc),
                "link_name": infer_link_name(name),
                "sample_label_values": sample_label_values,
                "recommended_next_action": "优先使用 runtime labels，其次回退到 query/generatorURL 静态提取",
            })

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_DIR / "layer1_alert_families.enhanced.yaml", "w", encoding="utf-8") as f:
        yaml.safe_dump(layer1, f, allow_unicode=True, sort_keys=False)
    with open(OUT_DIR / "layer2_classifier_mapping.enhanced.yaml", "w", encoding="utf-8") as f:
        yaml.safe_dump({"layer2_classifier_mapping": layer2}, f, allow_unicode=True, sort_keys=False)
    with open(OUT_DIR / "layer3_context_enrichment.enhanced.yaml", "w", encoding="utf-8") as f:
        yaml.safe_dump({"layer3_context_enrichment": layer3}, f, allow_unicode=True, sort_keys=False)
    with open(OUT_DIR / "device_inventory.normalized.yaml", "w", encoding="utf-8") as f:
        yaml.safe_dump({"devices": inventory_rows}, f, allow_unicode=True, sort_keys=False)
    with open(OUT_DIR / "target_lookup.yaml", "w", encoding="utf-8") as f:
        yaml.safe_dump({"targets": target_rows}, f, allow_unicode=True, sort_keys=False)

    print("generated:")
    print(OUT_DIR / "layer1_alert_families.enhanced.yaml")
    print(OUT_DIR / "layer2_classifier_mapping.enhanced.yaml")
    print(OUT_DIR / "layer3_context_enrichment.enhanced.yaml")
    print(OUT_DIR / "device_inventory.normalized.yaml")
    print(OUT_DIR / "target_lookup.yaml")

if __name__ == "__main__":
    build()
